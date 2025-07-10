create excel using openpy 



import openpyxl
import os


1. create the workbook  
 ==> wb = openpyxl.Workbook()

2. A workbook is always created with at least one worksheet. You can get it by using the Workbook.active property:

==> ws = wb.active

3. You can create new worksheets using the Workbook.create_sheet() method:

wb = wb.create_sheet("Mysheet") # insert at the end (default)
# or
wb = wb.create_sheet("Mysheet", 0) # insert at first position
# or
wb = wb.create_sheet("Mysheet", -1) # insert at the penultimate position

4.  You can change this name at any time with the Worksheet.title property:

==> wb.title = "New Title"

5 Once you gave a worksheet a name, you can get it as a key of the workbook:

==> wb = wb["New Title"]

6. You can review the names of all worksheets of the workbook with the Workbook.sheetname attribute

==> print(wb.sheetnames)
 
7 You can loop through worksheets
==>for sheet in wb:
    print(sheet.title)

8 You can create copies of worksheets within a single workbook: Workbook.copy_worksheet() method:

==> source = wb.active
    target = wb.copy_worksheet(source) 

Playing with data

1. Accessing one cell 
 ==> c = wb['A4']

2. This will return the cell at A4, or create one if it does not exist yet. Values can be directly assigned:
 ==> wb['A4'] = 4

3. There is also the Worksheet.cell() method. This provides access to cells using row and column notation:
 ==> d = ws.cell(row=4, column=2, value=10) 

 Note When a worksheet is created in memory, it contains no cells. They are created when first accessed.

Warning Because of this feature, scrolling through cells instead of accessing them directly will create them all in memory, even if you don’t assign them a value.
Something like

== > for x in range(1,101):
       for y in range(1,101):
           wb.cell(row=x, column=y)


4. Accessing many cells Ranges of cells can be accessed using slicing:
  ==> cell_range = wb['A1':'C2']

5. Ranges of rows or columns can be obtained similarly:

 ==>  colC = wb['C']
     col_range = wb['C:D']
     row10 = wb[10]
     row_range = wb[5:10]


6. You can also use the Worksheet.iter_rows() method:
 ==> for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
  		 for cell in row: 
			print(cell)

7. Likewise the Worksheet.iter_cols() method will return columns:
   ==> for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
        for cell in col:
          print(cell)

8. Note For performance reasons the Worksheet.iter_cols() method is not available in read-only mode.
   If you need to iterate through all the rows or columns of a file, you can instead use the Worksheet.rows property:
   ==> wb = wb.active
	wb['C9'] = 'hello world'
	tuple(wb.rows)


9. or the Worksheet.columns property:
    ==>tuple(wb.columns)

Note For performance reasons the Worksheet.columns property is not available in read-only mode.

10. Values only
If you just want the values from a worksheet you can use the Worksheet.values property. This iterates over all the rows in a worksheet but returns just the cell values:

==>for row in ws.values:
    for value in row:
      print(value)

11. Both Worksheet.iter_rows() and Worksheet.iter_cols() can take the values_only parameter to return just the cell’s value:

==> for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
  	print(row)


Data storage

1.Once we have a Cell, we can assign it a value: 
 ==>c.value = 'hello, world'
   print(c.value)

2. The simplest and safest way to save a workbook is by using the Workbook.save() method of the Workbook object: 

==> wb = Workbook()
    wb.save('balances.xlsx')

3. Saving as a stream
If you want to save the file to a stream, e.g. when using a web application such as Pyramid, Flask or Django then you can simply provide a NamedTemporaryFile():

from tempfile import NamedTemporaryFile
from openpyxl import Workbook
wb = Workbook()
with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()



Loading from a file

1. You can use the openpyxl.load_workbook() to open an existing workbook:

from openpyxl import load_workbook
wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)


# adds

#merge cells 
sheetname.merge_cells("A2:H2")

# Center-align the content of the cell
sheetname['A2'].alignment = Alignment(horizontal='center', vertical='center')

# Apply bold formatting to the cell
bold_font = Font(bold=True)
sheetname['A2'].font = bold_font

#cell_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # red

# Add themed text to the column
data_font = Font(color="FF0000")

